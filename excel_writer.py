"""
excel_writer.py
===============
Write the final formatted Excel output.

Column layout (matches user-provided template):
  A  → PDF_Title
  B  → Introduction / Preamble
  C  → Vision / Objectives / Goals
  D  → Key Policy Provisions / Measures
  E  → Implementation Plan / Action Points
  F  → Target Groups / Beneficiaries
  G  → Institutional Framework / Governance
  H  → Monitoring, Evaluation & Review
  I  → Financial Provisions / Resource Allocation

Features:
  • Professional header styling
  • Alternating row colours
  • Freeze panes (B2)
  • Auto column widths
  • Summary statistics sheet
  • Colour-coded error / skipped / pending rows
"""

import logging
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config as cfg

logger = logging.getLogger("pdf_pipeline")

# ─── Styles ──────────────────────────────────────────────────────────────────
_H_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy   — header row
_ALT_FILL = PatternFill("solid", fgColor="EBF3FB")   # light blue  — alt data rows
_WHT_FILL = PatternFill("solid", fgColor="FFFFFF")   # white       — data rows
_ERR_FILL = PatternFill("solid", fgColor="FFE0E0")   # light red   — errors
_SKP_FILL = PatternFill("solid", fgColor="FFF8E1")   # light amber — skipped
_PND_FILL = PatternFill("solid", fgColor="F5F5F5")   # light grey  — pending

_H_FONT   = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
_TL_FONT  = Font(name="Arial", bold=True,  color="1F3864", size=9)
_CELL_FNT = Font(name="Arial",             color="000000", size=9)
_ERR_FONT = Font(name="Arial", italic=True, color="CC0000", size=9)
_SKP_FONT = Font(name="Arial", italic=True, color="B8860B", size=9)
_PND_FONT = Font(name="Arial", italic=True, color="808080", size=9)

_WRAP     = Alignment(wrap_text=True, vertical="top",    horizontal="left")
_CENTER   = Alignment(wrap_text=True, vertical="center", horizontal="center")

def _thin(color="B8CCE4"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────

def write_excel(
    output_path: str,
    all_items:   List[Dict[str, str]],
    completed:   Dict[str, Dict[str, str]],
    failed:      Dict[str, str],
    skipped:     List[str],
    run_metadata: Dict[str, Any],
) -> None:
    """
    Write the final Excel file.

    Parameters
    ----------
    output_path   : Destination .xlsx file path.
    all_items     : Ordered list of {title, path} dicts (preserves row order).
    completed     : {title → {section_key → text}} from checkpoint.
    failed        : {title → error_message} from checkpoint.
    skipped       : [title, ...] — PDFs where no text could be extracted.
    run_metadata  : {started_at, last_updated, ...} for summary sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Policy Summaries"

    # ── Header row ─────────────────────────────────────────────────────────
    headers = ["PDF_Title"] + [s["name"] for s in cfg.FRAMEWORK_SECTIONS]
    for col_i, header in enumerate(headers, start=1):
        c           = ws.cell(row=1, column=col_i, value=header)
        c.fill      = _H_FILL
        c.font      = _H_FONT
        c.alignment = _CENTER
        c.border    = _thin()
    ws.row_dimensions[1].height = 32

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_offset, item in enumerate(all_items):
        excel_row  = row_offset + 2          # row 1 = header
        title      = item["title"]
        row_fill   = _ALT_FILL if row_offset % 2 == 0 else _WHT_FILL
        sections   = completed.get(title)
        error_msg  = failed.get(title)
        is_skipped = title in skipped

        # --- Title cell ---
        tc            = ws.cell(row=excel_row, column=1, value=title)
        tc.fill       = row_fill
        tc.font       = Font(name="Arial", bold=True, size=9)
        tc.alignment  = _WRAP
        tc.border     = _thin()

        if sections:
            # Successfully processed
            for col_i, sec in enumerate(cfg.FRAMEWORK_SECTIONS, start=2):
                text        = sections.get(sec["key"], "N/A")
                c           = ws.cell(row=excel_row, column=col_i, value=text)
                c.fill      = row_fill
                c.font      = _CELL_FNT
                c.alignment = _WRAP
                c.border    = _thin()

        elif error_msg:
            # Failed
            ec = ws.cell(row=excel_row, column=2,
                         value=f"[ERROR] {error_msg[:500]}")
            ec.font  = _ERR_FONT
            ec.alignment = _WRAP
            for col_i in range(2, len(headers) + 1):
                ws.cell(row=excel_row, column=col_i).fill   = _ERR_FILL
                ws.cell(row=excel_row, column=col_i).border = _thin()
            tc.fill = _ERR_FILL

        elif is_skipped:
            # Skipped — no extractable text
            sc = ws.cell(row=excel_row, column=2,
                         value="[SKIPPED] Insufficient text could be extracted from this PDF.")
            sc.font  = _SKP_FONT
            sc.alignment = _WRAP
            for col_i in range(2, len(headers) + 1):
                ws.cell(row=excel_row, column=col_i).fill   = _SKP_FILL
                ws.cell(row=excel_row, column=col_i).border = _thin()
            tc.fill = _SKP_FILL

        else:
            # Not yet processed
            pc = ws.cell(row=excel_row, column=2, value="[NOT PROCESSED]")
            pc.font  = _PND_FONT
            pc.alignment = _WRAP
            for col_i in range(2, len(headers) + 1):
                ws.cell(row=excel_row, column=col_i).fill   = _PND_FILL
                ws.cell(row=excel_row, column=col_i).border = _thin()
            tc.fill = _PND_FILL

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    for col_i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 48

    # ── Freeze panes ───────────────────────────────────────────────────────
    ws.freeze_panes = "B2"

    # ── Summary Statistics sheet ───────────────────────────────────────────
    ws2 = wb.create_sheet("Summary Statistics")
    stats_rows = [
        ("Metric",                  "Value"),
        ("Total PDFs",              len(all_items)),
        ("Successfully Processed",  len(completed)),
        ("Failed",                  len(failed)),
        ("Skipped (no text)",       len(skipped)),
        ("Pending / Not Processed", len(all_items) - len(completed) - len(failed) - len(skipped)),
        ("Active LLM Provider",     cfg.ACTIVE_LLM.title()),
        ("Run Started",             run_metadata.get("started_at", "N/A")),
        ("Last Updated",            run_metadata.get("last_updated", "N/A")),
        ("OCR Output Directory",    cfg.OCR_OUTPUT_DIR),
        ("Output Excel",            output_path),
    ]
    for r, (label, val) in enumerate(stats_rows, start=1):
        is_header = r == 1
        lc = ws2.cell(row=r, column=1, value=label)
        vc = ws2.cell(row=r, column=2, value=val)
        lc.font = Font(name="Arial", bold=True,  size=10 if is_header else 9,
                       color="FFFFFF" if is_header else "000000")
        vc.font = Font(name="Arial",              size=10 if is_header else 9)
        if is_header:
            lc.fill = vc.fill = _H_FILL
        lc.alignment = vc.alignment = Alignment(vertical="center")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50
    ws2.row_dimensions[1].height = 24

    wb.save(output_path)
    logger.info(f"Excel written → {output_path}")
