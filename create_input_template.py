"""
create_input_template.py
========================
Generates input_template.xlsx matching the user-provided format:
  Col A: PDF_Title
  Col B: PDF_Path
  (Additional 8 output columns are filled by the pipeline itself)

Run once before using the pipeline:
    python create_input_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config as cfg

H_FILL  = PatternFill("solid", fgColor="1F3864")
H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
EX_FILL = PatternFill("solid", fgColor="EBF3FB")
CELL_FN = Font(name="Arial", size=10)
THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
WRAP    = Alignment(wrap_text=True, vertical="top")
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_template(path: str = "input_template.xlsx"):
    wb = Workbook()

    # ── Sheet 1: PDF_List (input) ──────────────────────────────────────
    ws = wb.active
    ws.title = "PDF_List"

    # Header row: PDF_Title + 8 output section columns
    headers = ["PDF_Title", "PDF_Path"] + [s["name"] for s in cfg.FRAMEWORK_SECTIONS]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_i, value=h)
        c.fill      = H_FILL
        c.font      = H_FONT
        c.alignment = CENTER
        c.border    = THIN

    # Example data rows
    examples = [
        ("National Health Policy 2017",    "./pdfs/national_health_policy.pdf"),
        ("Digital India Programme",        "./pdfs/digital_india.pdf"),
        ("National Education Policy 2020", "/absolute/path/nep_2020.pdf"),
        ("PM Kisan Scheme",                "C:\\Users\\user\\pdfs\\pm_kisan.pdf"),
    ]
    for row_i, (title, pdf_path) in enumerate(examples, start=2):
        fill = EX_FILL if row_i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_i, val in enumerate([title, pdf_path], 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill      = fill
            c.font      = CELL_FN
            c.alignment = WRAP
            c.border    = THIN
        # Leave section columns blank — pipeline fills them
        for col_i in range(3, len(headers) + 1):
            ws.cell(row=row_i, column=col_i).fill   = fill
            ws.cell(row=row_i, column=col_i).border = THIN

    # Column widths
    ws.column_dimensions["A"].width = 40   # PDF_Title
    ws.column_dimensions["B"].width = 55   # PDF_Path
    for col_i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 48
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Sheet 2: Instructions ──────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    rows = [
        ("COLUMN",    "DESCRIPTION",                                               "REQUIRED"),
        ("PDF_Title", "Display name of the PDF. Used as row label in output.",     "YES"),
        ("PDF_Path",  "Full OR relative path to the PDF file. "
                      "Relative paths are resolved from the working directory "
                      "or --pdf-dir argument.",                                    "YES"),
        ("Sections",  "Columns C–J are filled automatically by the pipeline. "
                      "Leave them blank in the input file.",                       "NO"),
        ("", "", ""),
        ("TIPS", "", ""),
        ("•", "One row per PDF. Keep PDF_Title unique.", ""),
        ("•", "Use forward slashes for paths on Linux/macOS.", ""),
        ("•", "On Windows use backslashes or raw strings (C:\\path\\file.pdf).", ""),
        ("•", "The pipeline silently skips rows where the PDF file is not found.", ""),
        ("•", "Delete example rows and add your own PDFs.", ""),
        ("", "", ""),
        ("LLM PROVIDERS", "", ""),
        ("anthropic", "Claude (Anthropic) — set ANTHROPIC_API_KEY in .env",        ""),
        ("openai",    "GPT-4o (OpenAI)   — set OPENAI_API_KEY in .env",            ""),
        ("gemini",    "Gemini (Google)   — set GEMINI_API_KEY in .env",            ""),
        ("ollama",    "Ollama (local)    — no API key needed; ollama must run",     ""),
    ]
    hdr_font = Font(name="Arial", bold=True, size=10)
    cell_fn2 = Font(name="Arial", size=9)
    for r, (a, b, c_val) in enumerate(rows, start=1):
        is_hdr = r == 1 or a in ("TIPS", "LLM PROVIDERS")
        ws2.cell(row=r, column=1, value=a).font = hdr_font if is_hdr else cell_fn2
        ws2.cell(row=r, column=2, value=b).font = cell_fn2
        ws2.cell(row=r, column=3, value=c_val).font = Font(name="Arial", bold=(c_val == "YES"), size=9)
        if r == 1:
            for col_i in range(1, 4):
                ws2.cell(row=1, column=col_i).fill = H_FILL
                ws2.cell(row=1, column=col_i).font = H_FONT
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 12

    wb.save(path)
    print(f"Template created: {path}")
    print("Edit the PDF_List sheet — add your PDF titles and paths.")
    print("Then run:  python main.py --input input_template.xlsx")


if __name__ == "__main__":
    make_template()
